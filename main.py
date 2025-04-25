import openpyxl
import os , shutil
from os import listdir
import re
from openpyxl.styles import PatternFill, Alignment, Font
import datetime
from utils import getProccessedPdfs
from extract_image import extract_DPE_from_image
from tqdm import tqdm
DATE_REGEXP = "(\d+)/(\d+)/(\d+)"

CLEANUP_TXT = True

COLUMNS = {
    "Code de l’ESI": "A",
    "Code Groupe": "D",
    "Code de l’ESI2": "E",
    "Consommation (DPE) kWh/m2/an": "O",
    "Emissions (DPE) Kg CO2/m2/an": "P",
    "Lettre Consommation": "Q",
    "Emissions de gaz à effet de serre (DPE) Kg CO2/m2/an": "R",
    "Lettre Emissions": "S",
    "Etabli le": "T",
    "Valable jusqu’au": "U",
    "N°ADEME": "V",
    "Surface habitable (DPE)m2": "W",
    "Année de construction (DPE)": "X",
    "Estimation des coûts annuels d’énergie du logement": "Y",
    "Année de référence des prix énergétiques du diagnostic énergétique à l’origine de l’estimation": "Z",
    "Année de référence des prix énergétiques du diagnostic énergétique à l’origine de l’estimation  - Année": "AA",
    "Surface habitable de l’immeuble (DPE) m2": "AB",
    "Système de ventilation en place (DPE)": "AC",
    "Chauffage (DPE)": "AD",
    "Chauffage - Equipement (DPE)": "AE",
    "Eau Chaude (DPE)": "AF",
    "Eau Chaude sanitaire - équipement (DPE)": "AG",
    "Refroidissement (DPE)": "AH",
    "Eclairage (DPE)": "AI",
    "Auxiliares (DPE)": "AJ",
    "Volets / brise soleil": "AK",
    "Production d'énergie renouvelable": "AL",
    "Confort d'été (caractéristiques du logement améliorant le confort d'été)": "AM",
}


COLORS = {
    "A":"078E34",
    "B":"37AB32",
    "C":"A0CD63",
    "D":"ffff00",
    "E":"fabf8f",
    "F":"e26b09",
    "G":"c00000"
}




def determiner_classe_consommation(consommation):
    if consommation <= 70:
        return "A"
    elif consommation <= 110:
        return "B"
    elif consommation <= 180:
        return "C"
    elif consommation <= 250:
        return "D"
    elif consommation <= 330:
        return "E"
    elif consommation <= 420:
        return "F"
    else:
        return "G"


def determiner_classe_emission(emissions):
    if emissions <= 5:
        return "A"
    elif emissions <= 11:
        return "B"
    elif emissions <= 30:
        return "C"
    elif emissions <= 50:
        return "D"
    elif emissions <= 70:
        return "E"
    elif emissions <= 100:
        return "F"
    else:
        return "G"


class PDFEntry():
    def __init__(self, title, text, workbook) -> None:
        self.title = title
        self.text = text
        self.pages = text.split("=SPLIT=")
        self.data = {}
        self.text_pyocr=""
        self.workbook = workbook
        self.is_valid = True
        self.error = ""
    
    def set_esi(self):
        self.data[f'{COLUMNS["Code de l’ESI"]}'] = self.title.split("_")[0]
        self.data[f'{COLUMNS["Code de l’ESI2"]}'] = self.title.split("_")[0]
    
    def set_groupe(self):
        self.data[f'{COLUMNS["Code Groupe"]}'] = self.data[f'{COLUMNS["Code de l’ESI"]}'].split("H")[0]

    def set_consommation(self):
        try:
            
            result = re.findall("émissions\n+(\d+) *(\||,|\(|\)|\/\|) *(\d+).*", self.pages[-1])
            dpe = result[0][0]
            self.data[f'{COLUMNS["Consommation (DPE) kWh/m2/an"]}'] = dpe
        except Exception as Err:
            print(self.title)
            self.is_valid = False
            self.error += "Consommation, "
            print("Consommation Error", Err)

    def set_emission(self):
        try:
            res = re.findall(
                "émissions\n+(\d+) *(\||,|\(|\)|\/\|) *(\d+).*", self.pages[-1])[0][-1]
            self.data[f'{COLUMNS["Emissions (DPE) Kg CO2/m2/an"]}'] = res
        except Exception as Err:
            print(self.title)
            self.is_valid = False
            self.error += "Emissions, "
            print("Emissions (DPE) Kg CO2/m2/an ", Err)

    def set_emission_gaz(self):
        try:
            self.data[f'{COLUMNS["Emissions de gaz à effet de serre (DPE) Kg CO2/m2/an"]}'] = self.data[
                f'{COLUMNS["Emissions (DPE) Kg CO2/m2/an"]}']
        except Exception as Err:
            print(self.title)
            self.is_valid = False
            self.error += "Emission gaz, "
            print("Emission gaz Error", Err)

    def set_lettreConsommation(self):
        try:
            self.data[f'{COLUMNS["Lettre Consommation"]}'] = determiner_classe_consommation(
                int(self.data[f'{COLUMNS["Consommation (DPE) kWh/m2/an"]}']))
        except Exception as Err:
            print(self.title)
            self.is_valid = False
            self.error += "Lettre Consommation, "
            print("Lettre Consommation", Err)

    def set_lettreEmission(self):
        try:
            self.data[f'{COLUMNS["Lettre Emissions"]}'] = determiner_classe_emission(
                int(self.data[f'{COLUMNS["Emissions (DPE) Kg CO2/m2/an"]}']))
        except Exception as Err:
            print(self.title)
            self.is_valid = False
            self.error += "Lettre Emission, "
            print("lettre emission Error", Err)

    def set_Etablile(self):
        try:
            EtabliLe = re.findall("Etabli le : (\d+)/(\d+)/(\d+)", self.pages[0])
            Etabli = '/'.join(EtabliLe[0])
            self.data[f'{COLUMNS["Etabli le"]}'] = Etabli
        except Exception as Err:
            print(self.title)
            self.is_valid = False
            self.error += "Etabli le, "
            print("Etabli le :", Err)

    def set_NADEME(self):
        try:
            NADEME = re.findall("ADEME : ([A-Z0-9]+)", self.pages[0])
            NADEME = NADEME[0]
            self.data[f'{COLUMNS["N°ADEME"]}'] = NADEME
        except Exception as Err:
            print(self.title)
            self.is_valid = False
            self.error += "N°ADEME, "
            print("NADEME Error", Err)

    def set_surfaceHabitable(self):
        try:
            surfaceHabitable = re.findall(
                "Surface habitable : (\d+,*\d*)", self.pages[0])
            self.data[f'{COLUMNS["Surface habitable (DPE)m2"]}'] = surfaceHabitable[0]
        except Exception as Err:
            print(self.title)
            self.is_valid = False
            self.error += "Surface habitable, "
            print("Surface habitable", Err)

    def set_Valablej(self):
        try:
            Valablej = re.findall(
                "Valable jusqu’au : (\d+)/(\d+)/(\d+)", self.pages[0])
            Valable = '/'.join(Valablej[0])
            self.data[f'{COLUMNS["Valable jusqu’au"]}'] = Valable
        except Exception as Err:
            print(self.title)
            self.is_valid = False
            self.error += "Valable jusqu'au, "
            print("valable Error", Err)

    def set_anneeconstruction(self):
        try:
            anneeconstruction = re.findall(
                "Année de construction : (\d+)", self.pages[0])
            anneeconstruction = anneeconstruction[0]
            self.data[f'{COLUMNS["Année de construction (DPE)"]}'] = anneeconstruction
        except Exception as Err:
            print(self.title)
            self.is_valid = False
            self.error += "Année de construction, "
            print("anneeconstruction Error", Err)

    def set_estimation(self):
        try:
            estimation = re.findall(
                "([0-9 ]+)€[^0-9]*([0-9 ]+)", self.pages[0])
            estimation = f"entre {estimation[0][0].strip()} € et {estimation[0][1].strip()} € par an"
            self.data[f'{COLUMNS["Estimation des coûts annuels d’énergie du logement"]}'] = estimation
        except Exception as Err:
            print(self.title)
            self.is_valid = False
            self.error += "Estimation des coûts annuels, "
            print("estimation Error", Err)

    def set_surfaceHabitableImmeuble(self):
        try:
            temp_pages = "\n".join(self.pages[1:])
            SurfaceHabitableImmeuble = re.findall(
                "Surface habitable de l\'immeuble (.*) mesuré (\d+,*\d+) m", temp_pages)
            if(len(SurfaceHabitableImmeuble) == 0):
                SurfaceHabitableImmeuble = re.findall("Surface habitable de l\'immeuble (.*) (\d+,*\d+) m", temp_pages)
            #print(SurfaceHabitableImmeuble)
            self.data[f'{COLUMNS["""Surface habitable de l’immeuble (DPE) m2"""]}'] = SurfaceHabitableImmeuble[0][1]
        except Exception as Err:
            print(self.title)
            self.is_valid = False
            self.error += "Surface habitable immeuble, "
            print("surface habitable immeuble Error", Err)

    def set_prixmoyen(self):
        try:
            prixmoyen = re.findall(
                "Prix moyens des énergies indexés .* \(abonnements compris\)", self.pages[0])
            self.data[f'{COLUMNS["Année de référence des prix énergétiques du diagnostic énergétique à l’origine de l’estimation"]}'] = prixmoyen[0]
        except Exception as Err:
            print(self.title)
            self.is_valid = False
            self.error += "Prix moyen, "
            print("estimation Error", Err)

    def set_anneeprixmoyen(self):
        try:
            anneeprixmoyen = re.findall('indexés .* (\d+) \(abonnements compris\)',
                                        self.data[f'{COLUMNS["Année de référence des prix énergétiques du diagnostic énergétique à l’origine de l’estimation"]}'])
            self.data[f'{COLUMNS["Année de référence des prix énergétiques du diagnostic énergétique à l’origine de l’estimation  - Année"]}'] = anneeprixmoyen[0]
        except Exception as Err:
            print(self.title)
            self.is_valid = False
            self.error = "Année prix moyen, "
            print("indexés", Err)

    def get_data(self):
        return self.data

    def __str__(self) -> str:
        res = ""
        for key, value in self.data.items():
            res += f"{key} : {value} \n"
        return res


def scrape(wb):
    folder = "txt"
    if not os.path.exists(folder):
        #print("there is not txt folder to scape ! exiting..")
        return
    ws = wb.active
    ft1 = Font(name='Arial', size=8)
    ft2 = Font(name='Arial', size=12, bold=True)
    row = ws.max_row - 1
    PROCESSED_PDFS = getProccessedPdfs(ws)
    errorsFile = open("errors.txt","a+",encoding="utf-8")
    for file in listdir(folder):
        if(file.split("_")[0] in PROCESSED_PDFS):
            #print(f"this file : {file.split('_')[0]} is already in this excel file , aborting scraping")
            continue
        with open(os.path.join(folder,file),encoding="utf-8") as f:
            str = f.read()
            test = PDFEntry(file, str, ws)
            test.set_esi()
            test.set_groupe()
            # extraction image pour consommation et émissions
            test.set_consommation()
            test.set_emission()
            test.set_lettreConsommation()
            test.set_emission_gaz()
            test.set_lettreEmission()
            test.set_Etablile()
            test.set_Valablej()
            test.set_NADEME()
            test.set_surfaceHabitable()
            test.set_anneeconstruction()
            test.set_estimation()
            test.set_prixmoyen()
            test.set_anneeprixmoyen()
            #test.set_surfaceHabitableImmeuble()
            data = test.get_data()
            for key, value in data.items():
                    rd = ws.row_dimensions[row]
                    rd.height = 34.5
                    ws[f"{key}{row}"] = value
                    ws[f"{key}{row}"].font = ft1
                    ws[f"{key}{row}"].alignment = Alignment(horizontal='center',vertical='center')
                    if(key == "O" or key == "P" or key=="R" or key=="W" or key == "X" or key == "AA" or key == "AB"):
                        new_value = value.replace(",",".")
                        ws[f"{key}{row}"] = float(new_value)
                    if(key == "T" or key == "U" or key == "V" or key == "Y"):
                        ws[f"{key}{row}"].alignment = Alignment(vertical='center')
                    if(key == "Z"):
                        ws[f"{key}{row}"].alignment = Alignment(horizontal='left',vertical='center',wrap_text=True)
                    if(key == "Q"):
                        color = COLORS[f'{value.upper()}']
                        ws[f"{key}{row}"].fill = PatternFill(fill_type='solid', bgColor=color, start_color=color, end_color=color)
                        ws[f"{key}{row}"].font = ft2
                    if(key == "Y" or key == "AA"):
                        ws[f"{key}{row}"].fill = PatternFill(fill_type='solid',  bgColor='BFDDE7',start_color='BFDDE7', end_color='BFDDE7')
                    if(key == "T" or key == "U"):
                        ws[f"{key}{row}"] = datetime.datetime.strptime(value, "%d/%m/%Y")
                        ws[f"{key}{row}"].number_format = 'DD/MM/YYYY'
            if(not test.is_valid):
                errorsFile.write(f"{file}:{test.error}\n")
            row += 1
    errorsFile.close()       
    wb.save("excel.xlsx")
    print("done, saving to excel , deleting txt folder")
    if CLEANUP_TXT:
        for files in os.listdir(folder):
            path = os.path.join(folder, files)
            try:
                shutil.rmtree(path)
            except OSError:
                os.remove(path)
        os.removedirs(folder)
    wb.close()


if __name__ == "__main__":
    wb = openpyxl.load_workbook('test.xlsx')
    ws = wb.active
    row = ws.max_row - 1
    folder = "./txt"
    for file in listdir(folder):
        with open(os.path.join(folder,file),encoding="utf-8") as f:
            str = f.read()
            test = PDFEntry(file, str, ws)
            test.set_consommation()
            test.set_emission()
            test.set_lettreConsommation()
            test.set_emission_gaz()
            test.set_lettreEmission()
            test.set_Etablile()
            test.set_Valablej()
            test.set_NADEME()
            test.set_surfaceHabitable()
            test.set_anneeconstruction()
            test.set_estimation()
            test.set_prixmoyen()
            test.set_anneeprixmoyen()
            #test.set_surfaceHabitableImmeuble()
            data = test.get_data()
            for key, value in data.items():
                ws[f"{key}{row}"] = value
            row += 1

    wb.save("test.xlsx")
